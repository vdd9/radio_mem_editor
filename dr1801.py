#!/usr/bin/python

import re
from common import *
import openpyxl
from openpyxl.utils import get_column_letter

# note:
# stangre values after 0x0001C6C0
# encryption 0x0001D7E0
# VFO sttings 0x0001DD00

class Zone:
    def __init__(self, id, block, offset, _):
        struct.pack_into('<H', block, offset+0x24, id)
        self.block = block
        self.offset = offset
        self.val_of_ = ByteValueHandler(block, offset, {
            'namesize':    0x20,
            'channels_nbr': 0x22
        })
            
    @property
    def channels_IDs(self):
        return [ struct.unpack('<H',self.block[self.offset+0x28+2*i:self.offset+0x2A+2*i])[0] for i in range(0,self.val_of_['channels_nbr']) ]
    @channels_IDs.setter
    def channels_IDs(self, new_list):
        self.val_of_["channels_nbr"] = len(new_list)
        for i in range(0,len(new_list)):
            struct.pack_into('<H',self.block,self.offset+0x28+2*i,new_list[i])

    @property
    def name(self):
        return estract_string_til_zero(self.block[self.offset:], 0x20)
    @name.setter
    def name(self, new_name):
        self.val_of_['namesize'] = len(new_name[:0x20])
        self.block[self.offset+0x20] = len(new_name[:0x20])
        self.block[self.offset:self.offset+0x20] = bytearray(new_name[:0x20].ljust(0x20,'\x00'), 'utf-8')
    
    def __str__(self):
        return f'{self.name}: {self.channels_IDs}'

class Channel:
    channels_size = 0x34
    def __init__(self, id, block, offset, offset_name):
        struct.pack_into('<H', block, offset, id)
        self.block = block
        self.offset = offset
        self.offset_name = offset_name
        self.is_ = BoolHandler(block, offset, {
            'high':     0x03,
            'DTMF':     0x16,
            'wide':     0x19,
            'autoscan': 0x1A,
            'lonework': 0x2F
        },{
            'dmr': (0x02,0x01,0x03)
        },{
            'talkaround':     (0x15,0x80),
            'directTDMA':     (0x15,0x02),
            'privateconfirm': (0x15,0x01)
        })
        self.val_of_ = ByteValueHandler(block, offset, {
            'talkgroup':    0x0C,
            'PTTstate':     0x0E,
            'colorcode':    0x10,
            'slot':         0x11,
            'encryption':   0x14,
            'alarm':        0x18,
            'autoscanlist': 0x1B,
            'PTTID':        0x28,
            'receivegroup': 0x2A
        })

    ## clear ##

    def clear(self):
        for k in self.is_.keys():
            self.is_[k] = False
        for k in self.val_of_.keys():
            self.val_of_[k] = 0
        self.txDcs = ''
        self.rxDcs = ''

    ## ID ##
        
    @property
    def id(self):
        return struct.unpack('<H',self.block[self.offset:self.offset+2])[0]

    ## NAME ##

    @property
    def name(self):
        return estract_string_til_zero(self.block[self.offset_name:], 0x14)
    @name.setter
    def name(self, new_name):
        self.block[self.offset_name:self.offset_name+0x14] = bytearray(new_name[:0x14].ljust(0x14,'\x00'), 'utf-8')

    ## DSIPLAY ##

    def __str__(self):
        return f'{self.name}  '+('~','')[self.is_['dmr']]+'('+('low','high')[self.is_['high']]+('-narow','-wide')[self.is_['wide']]+f')\n RX: {(self.rxFreq):0>9.5f}  {self.rxDcs}\n TX: {(self.txFreq):0>9.5f}  {self.txDcs}'
    def __repr__(self):
        return self.name.ljust(12," ")+" ".join([str(b).zfill(3) for b in self.block[self.offset:self.offset+self.channels_size]])\
        + "\n            ID "+str(self.id).ljust(4," ")+"|"+("AMA","DMR")[self.is_['dmr']]+"|"+("Low","Hi ")[self.is_['high']]+"| RX: "+f'{(self.rxFreq):0>9.5f} | TX: {(self.txFreq):0>9.5f}'+" |TG |   |PTT|   |Col|Slt|       |Enc| & |DTM|   |Sys|"\
        +("NFM","FM ")[self.is_['wide']]+"|autosca|R: "+self.rxDcs.ljust(12)+"|T: "+self.txDcs.ljust(12)+"|               |PID|   |RG |               |lon|"
    def dump(self):
        print("========")
        print(self.name)
        print("RX:", self.rxFreq, self.rxDcs)
        print("TX:", self.txFreq, self.txDcs)
        for k in self.is_.keys():
            print(f'{k}:', self.is_[k])
        for k in self.val_of_.keys():
            print(f'{k}:', str(self.val_of_[str(k)]))

    ## FREQENCY ##

    @property
    def txFreq(self):
        return self._freq(8)
    @txFreq.setter
    def txFreq(self, new_freq):
        self._set_freq(8, new_freq)

    @property
    def rxFreq(self):
        return self._freq(4)
    @rxFreq.setter
    def rxFreq(self, new_freq):
        self._set_freq(4, new_freq)

    def _freq(self, idx):
        idx += self.offset
        return struct.unpack('<I',self.block[idx:idx+4])[0]/1000000
    def _set_freq(self, idx, new_freq):
        idx += self.offset
        struct.pack_into('<I', self.block, idx, int(new_freq*1000000))

    ## DCS ##

    @property
    def txDcs(self):
        return self._dcs(32)
    @txDcs.setter
    def txDcs(self, new_dcs):
        self._set_dcs(32, new_dcs)

    @property
    def rxDcs(self):
        return self._dcs(28)
    @rxDcs.setter
    def rxDcs(self, new_dcs):
        self._set_dcs(28, new_dcs)
    
    def _dcs(self, idx):
        idx += self.offset
        if self.block[idx+2] == 0:
            return ''
        elif self.block[idx+2] == 1:
            return "CTCSS " + str(struct.unpack('<H',self.block[idx:idx+2])[0]/10) + "Hz"
        elif self.block[idx+2] == 2:
            return "DCS " + str(struct.unpack('<H',self.block[idx:idx+2])[0]) + ("N","I")[self.block[idx+3]]
        return "WTF"
    def _set_dcs(self, idx, new_dcs):
        idx += self.offset
        if new_dcs and len(new_dcs)>0:
            if new_dcs.startswith("D"):
                self.block[idx+2] = 2
                struct.pack_into('<H', self.block, idx, int(re.findall(r"(\d+)", new_dcs)[0]))
                self.block[idx+3] = new_dcs.endswith("I")
            else:
                self.block[idx+2] = 1
                struct.pack_into('<H', self.block, idx, int(float(re.findall(r"(?:\d*\.*\d+)", new_dcs)[0])*10))
        else:
            struct.pack_into('<I', self.block, idx, 0)

class DR1801():
    scanlist_blocs = []
    def __init__(self,file):
        self.file = file
        self.ba = bytearray(self.file.read())
        self.channels = binaryList(Channel,self.ba,0xA660,0x34,0xA65C,0x00017660,0x14)
        self.scanlist_number = struct.unpack('<H',self.ba[0x0000A338:0x0000A33A])[0]
        self.zones = binaryList(Zone,self.ba,0x0420,0x68,0x0418)
        # fill scanlist
        for i in range(0,self.scanlist_number):
            start_addr = 0x0000A33B+0x50*i
            self.scanlist_blocs.append(self.ba[start_addr:start_addr+0x50])

    def importxls(self, path='freqs.xlsx'):
        if path.endswith(".xlsx"):
            book = openpyxl.load_workbook(path)
            # Parse Channels
            freqs_sheet = book.worksheets[0]
            chan_ix = 0
            previous_freq=0.0
            self.channels.length = freqs_sheet.max_row
            for row in freqs_sheet.iter_rows():
                current_chan = self.channels[chan_ix]
                current_chan.clear()
                current_chan.name = row[0].value
                freq = str(row[1].value)
                if freq.startswith('+') or freq.startswith('-'):
                    current_chan.txFreq = previous_freq + float(freq)
                else:
                    current_chan.txFreq = float(freq)
                freq = row[2].value
                print(row)
                print(freq)
                if freq:
                    current_chan.rxFreq = current_chan.txFreq + float(freq)
                else:
                    current_chan.rxFreq = current_chan.txFreq
                if row[3].value:
                    current_chan.txDcs = str(row[3].value)
                if row[4].value:
                    current_chan.rxDcs = str(row[4].value)
                if row[5].value:
                    for boolname in row[5].value.split('/'):
                        current_chan.is_[boolname] = True
                if row[6].value:
                    for val in row[6].value.split('/'):
                        k, v = val.split(":")
                        current_chan.val_of_[k] = int(v)
                previous_freq=current_chan.txFreq
                chan_ix +=1
            # Parse Zones
            zones_sheet = book.worksheets[1]
            self.zones.length = zones_sheet.max_column
            for cx in range(1,zones_sheet.max_column+1):
                current_zone = self.zones[cx-1]
                current_zone.name = zones_sheet[get_column_letter(cx)+'1'].value
                current_zone.channels_IDs = [int(c[0].value[8:])-1 for c in zones_sheet[get_column_letter(cx)+'2:'+get_column_letter(cx)+'33'] if c[0].value]

        else:
            book = xlrd.open_workbook(path)
            # Parse Channels
            freqs_sheet = book.sheet_by_index(0)
            chan_ix = 0
            previous_freq=0.0
            self.channels.length = freqs_sheet.nrows
            for rx in range(freqs_sheet.nrows):
                current_chan = self.channels[chan_ix]
                current_chan.clear()
                current_chan.name = freqs_sheet.row(rx)[0].value
                freq = str(freqs_sheet.row(rx)[1].value)
                if freq.startswith('+') or freq.startswith('-'):
                    current_chan.txFreq = previous_freq + float(freq)
                else:
                    current_chan.txFreq = float(freq)
                freq = str(freqs_sheet.row(rx)[2].value)
                if freq == "":
                    current_chan.rxFreq = current_chan.txFreq
                else:
                    current_chan.rxFreq = current_chan.txFreq + float(freq)
                current_chan.txDcs = str(freqs_sheet.row(rx)[3].value)
                current_chan.rxDcs = str(freqs_sheet.row(rx)[4].value)
                if freqs_sheet.row(rx)[5].value != "":
                    for boolname in freqs_sheet.row(rx)[5].value.split('/'):
                        current_chan.is_[boolname] = True
                if freqs_sheet.row(rx)[6].value != "":
                    for val in freqs_sheet.row(rx)[6].value.split('/'):
                        k, v = val.split(":")
                        current_chan.val_of_[k] = int(v)
                previous_freq=current_chan.txFreq
                chan_ix +=1
            # Parse Zones
                print("zone parsing from xls not implemented.")
        

    def save(self):
        self.file.seek(0)
        self.file.write(self.ba)

    def writexls(self, path='freqs.xlsx'):
        if path.endswith(".xlsx"):
            book = openpyxl.Workbook()
            # Export Channels
            freqs_sheet = book.active
            freqs_sheet.title = "chan"
            freqs_sheet.column_dimensions['A'].width = 11
            freqs_sheet.column_dimensions['B'].width = 11
            freqs_sheet.column_dimensions['C'].width = 11
            freqs_sheet.column_dimensions['F'].width = 11
            previous_freq = 0.0
            row_id = 1
            for chan in self.channels:
                freqs_sheet.cell(row=row_id,column=1).value = chan.name
                if chan.txFreq - previous_freq > 0 and chan.txFreq - previous_freq < 0.06:  
                    freqs_sheet.cell(row=row_id,column=2).value = f'+{(chan.txFreq-previous_freq):.5f}'
                else:
                    freqs_sheet.cell(row=row_id,column=2).value = f'{chan.txFreq:0>9.5f}'
                if chan.rxFreq != chan.txFreq:
                    freqs_sheet.cell(row=row_id,column=3).value = f'{(chan.rxFreq-chan.txFreq):+.5f}'
                freqs_sheet.cell(row=row_id,column=4).value = f'{chan.txDcs}'
                freqs_sheet.cell(row=row_id,column=5).value = f'{chan.rxDcs}'
                freqs_sheet.cell(row=row_id,column=6).value = '/'.join([k for k,v in chan.is_.items() if v == True])
                freqs_sheet.cell(row=row_id,column=7).value = '/'.join([f'{k}:{v}' for k,v in chan.val_of_.items() if v != 0])
                previous_freq=chan.txFreq
                row_id += 1

            # Export Zones
            zones_sheet = book.create_sheet("zone")
            zone_idx = 1
            for zone in self.zones:
                zones_sheet.cell(row=1,column=zone_idx).value = zone.name
                row_idx = 2
                for idx in zone.channels_IDs:
                    zones_sheet.cell(row=row_idx,column=zone_idx).value = "=chan!$A"+str(idx+1)
                    row_idx += 1
                zone_idx += 1

        else:
            book = xlwt.Workbook()
            # Export Channels
            freqs_sheet = book.add_sheet("chan")
            previous_freq = 0.0
            row_id = 0
            for chan in self.channels:
                freqs_sheet.write(row_id,0,chan.name)
                if chan.txFreq - previous_freq > 0 and chan.txFreq - previous_freq < 0.06:  
                    freqs_sheet.write(row_id,1,f'+{(chan.txFreq-previous_freq):.5f}')
                else:
                    freqs_sheet.write(row_id,1,f'{chan.txFreq:0>9.5f}')
                if chan.rxFreq == chan.txFreq:
                    freqs_sheet.write(row_id,2,'')
                else:
                    freqs_sheet.write(row_id,2,f'{(chan.rxFreq-chan.txFreq):+.5f}')
                freqs_sheet.write(row_id,3,f'{chan.txDcs}')
                freqs_sheet.write(row_id,4,f'{chan.rxDcs}')
                freqs_sheet.write(row_id,5,'/'.join([k for k,v in chan.is_.items() if v == True]))
                freqs_sheet.write(row_id,6,'/'.join([f'{k}:{v}' for k,v in chan.val_of_.items() if v != 0]))
                previous_freq=chan.txFreq
                row_id += 1
            for i in range(6):
                freqs_sheet.col(i).width = 256 * 12
            # Export Zones
            zones_sheet = book.add_sheet("zone")
            zone_idx = 0
            for zone in self.zones:
                zones_sheet.write(0,zone_idx,zone.name)
                row_idx = 1
                for idx in zone.channels_IDs:
                    zones_sheet.write(row_idx,zone_idx,"=chan!$A"+str(idx+1))
                    row_idx += 1
                zone_idx += 1

        book.save(path)

    def __del__(self):
        self.file.close()
